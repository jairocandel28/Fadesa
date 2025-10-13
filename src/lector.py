## módulo sys y os para interactuar con el sistema operativo

import sys
import os

## módulo para leer excel

from openpyxl import load_workbook

## módulo para leer SQLite

import sqlite3



def leer_csv(archivo) -> list:

    try:
        with open(archivo, "r", encoding = "utf-8") as f:
            try:
                texto = f.read()
            except UnicodeDecodeError:
                print(f"Error: El archivo '{archivo}' parece estar corrupto o tiene una codificación inválida.")
                sys.exit(1)

    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo '{archivo}'.")
        sys.exit(1)

    if not texto.strip():
        print(f"Error: El archivo '{archivo}' está vacío.")
        sys.exit(1)

    texto_completo = []

    for linea in texto.splitlines():
        if linea:
            linea_sin_espacios = linea.strip()
            texto_completo.append(linea_sin_espacios)

    return texto_completo



def leer_excel(archivo) -> list:
    try:
        wb = load_workbook(filename=archivo, data_only=True)
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo '{archivo}'.")
        sys.exit(1)
    except Exception as e:
        print(f"Error al abrir el archivo Excel '{archivo}': {e}")
        sys.exit(1)

    hoja = wb.active

    if hoja.max_row == 0 or hoja.max_column == 0:
        print(f"Error: El archivo '{archivo}' está vacío.")
        sys.exit(1)

    texto_completo = []
    for fila in hoja.iter_rows(values_only=True):
        if any(fila):  # evita filas vacías
            valores = [str(celda).strip() if celda is not None else "" for celda in fila]
            texto_completo.append(valores)

    return texto_completo



def leer_sqlite(archivo, tabla=None) -> list:
    try:
        conn = sqlite3.connect(archivo)
        cursor = conn.cursor()
    except sqlite3.Error as e:
        print(f"Error al abrir la base de datos '{archivo}': {e}")
        sys.exit(1)

    if tabla is None:
        try:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tablas = [t[0] for t in cursor.fetchall()]
            if not tablas:
                print(f"Error: La base de datos '{archivo}' no contiene tablas.")
                sys.exit(1)
            tabla = tablas[0]  # Por defecto, tomamos la primera
            print(f"No se especificó tabla, usando la primera encontrada: '{tabla}'")
        except sqlite3.Error as e:
            print(f"Error al obtener las tablas: {e}")
            sys.exit(1)

    try:
        cursor.execute(f"SELECT * FROM {tabla}")
        filas = cursor.fetchall()
        columnas = [desc[0] for desc in cursor.description]
    except sqlite3.Error as e:
        print(f"Error al leer la tabla '{tabla}': {e}")
        sys.exit(1)

    conn.close()

    texto_completo = [columnas] + [list(map(str, fila)) for fila in filas]

    return texto_completo



def leer_archivo(archivo):
    _, extension = os.path.splitext(archivo)
    extension = extension.lower()

    funciones = {
        ".csv": leer_csv,
        ".xlsx": leer_excel,
        ".xls": leer_excel,
        ".db": leer_sqlite
    }

    if extension not in funciones:
        print(f"Error: Formato de archivo '{extension}' no soportado.")
        return None

    # Ejecuta la función correspondiente
    return funciones[extension](archivo)


if __name__ == "__main__":

    
    archivo = input("Introduce la ruta completa del archivo que quieres analizar: ")
    resultado = leer_archivo(archivo)

    for fila in resultado:
        print(fila)

